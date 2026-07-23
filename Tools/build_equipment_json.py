"""
스탯 설계 xlsx의 '무기 수치표' / '방어구 수치표' / '재화 설계' 시트를 읽어
Assets/JsonData/ItemData.json 과 EquipmentData.json 을 생성한다.

수치가 가안이라 계속 바뀌므로, 시트를 손으로 펴서 JSON에 옮기지 않고 매번 이 스크립트로 다시 만든다.
시트가 원본(source of truth)이고 JSON은 산출물이다. JSON을 직접 고치면 다음 실행에서 덮어써진다.

사용법:
    pip install openpyxl
    python Tools/build_equipment_json.py "C:/path/to/stat_design.xlsx"

시트에 없는 정보(아이템 이름, 아이콘 주소)는 규칙으로 생성한다. 기획 확정 후 이름을 바꾸려면
NAME_TEMPLATE / ICON_TEMPLATE 을 수정하거나, 별도 이름 시트를 추가하고 이 스크립트를 확장한다.
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다:  pip install openpyxl")


# ── 설정 ────────────────────────────────────────────────────────────

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(REPO_ROOT, "Assets", "JsonData")

# 시트의 한글 등급명 → C# ItemGrade
GRADES = [("노말", "Normal"), ("매직", "Magic"), ("레어", "Rare"), ("유물", "Relic")]

# C# enum 값 → Index 자릿수. 값을 바꾸면 기존 Index가 전부 어긋나므로 건드리지 않는다.
CATEGORY_CODE = {"Weapon": 1, "Armor": 2}
SLOT_CODE = {"Sword": 1, "Gun": 2, "Top": 1, "Bottom": 2, "Gloves": 3, "Shoes": 4}

# 무기는 검/총 2벌, 방어구는 4파츠로 같은 수치를 복제한다.
WEAPON_VARIANTS = [("Sword", "검"), ("Gun", "총")]
ARMOR_VARIANTS = [("Top", "상의"), ("Bottom", "하의"), ("Gloves", "장갑"), ("Shoes", "신발")]

# 등급별 옵션 개수. 아래 OPTION_COUNT_OVERRIDE 로 예외를 덮는다.
OPTION_COUNT = {"Normal": 0, "Magic": 1, "Rare": 2, "Relic": 4}

# 스토리 클리어 보상인 Lv30 유물 무기는 유물인데 "옵션 없음"이다(기획서 명시).
# (카테고리, 등급, 레벨) → 옵션 개수
OPTION_COUNT_OVERRIDE = {("Weapon", "Relic", 30): 0}

SELL_PRICE_RATE = 0.2   # 판매가 = 구매가 × 20% (등급·레벨 무관 동일 비율)

NAME_TEMPLATE = "{grade_ko} {variant_ko} Lv{level}"
ICON_TEMPLATE = "Icon_{category}_{slot}_{grade}"


# ── 시트 파싱 ───────────────────────────────────────────────────────

def cell_number(value):
    """숫자 셀만 float으로 돌려준다. '-' 나 빈 칸은 None(해당 조합의 아이템 없음)."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace(",", "")
    if text in ("", "-", "—"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def find_row(ws, col, keyword, start=1, exact=False):
    """
    지정 열에서 keyword에 해당하는 첫 행 번호. 행이 밀려도 앵커로 찾기 위한 헬퍼.

    exact=True는 헤더 행을 찾을 때 쓴다. '레벨'을 부분 일치로 찾으면
    '※ 드랍 구간: ... 1~25레벨' 같은 주석 행이 먼저 걸린다.
    """
    for r in range(start, ws.max_row + 1):
        value = ws.cell(row=r, column=col).value
        if value is None:
            continue

        text = str(value).strip()
        if (text == keyword) if exact else (keyword in text):
            return r

    return None


def read_grade_columns(ws, header_row):
    """
    헤더 행에서 '노말/매직/레어/유물' 열 위치를 찾는다.

    등급명으로 시작하는지만 보면 '레어 4파츠+자체', '레어 판매가 (×20%)' 같은
    파생 열까지 걸려서 원본 열을 덮어쓴다(실제로 방어구 레어 기준값이 4,847로,
    판매가가 20%를 두 번 곱한 값으로 나왔다).
    그래서 등급명 뒤에 남는 것이 '(×1.25)' 같은 괄호 꼬리표뿐인 열만 받는다.
    """
    columns = {}
    for c in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=c).value
        if header is None:
            continue

        text = str(header).strip()
        for grade_ko, grade_en in GRADES:
            if not text.startswith(grade_ko):
                continue

            rest = text[len(grade_ko):].strip()
            if rest == "" or rest.startswith("("):
                columns.setdefault(grade_en, c)
            break

    return columns


def read_stat_table(ws):
    """수치표 시트 → { (등급, 레벨): 기준값 }"""
    header_row = find_row(ws, 2, "레벨", exact=True)
    if header_row is None:
        sys.exit(f"'{ws.title}' 시트에서 '레벨' 헤더를 찾지 못했습니다.")

    grade_columns = read_grade_columns(ws, header_row)
    if not grade_columns:
        sys.exit(f"'{ws.title}' 시트에서 등급 열을 찾지 못했습니다.")

    table = {}
    for r in range(header_row + 1, ws.max_row + 1):
        level = cell_number(ws.cell(row=r, column=2).value)
        if level is None:
            break   # 데이터 블록 끝

        for grade, column in grade_columns.items():
            value = cell_number(ws.cell(row=r, column=column).value)
            if value is not None:
                table[(grade, int(level))] = int(round(value))

    return table


def read_shop_table(ws, anchor):
    """상점 가격 블록 → { (등급, 레벨): 구매가 }. 정의되지 않은 구간은 아예 키가 없다."""
    anchor_row = find_row(ws, 2, anchor)
    if anchor_row is None:
        print(f"  [경고] '{anchor}' 블록을 찾지 못했습니다. 해당 판매가는 0이 됩니다.")
        return {}

    header_row = find_row(ws, 2, "레벨", start=anchor_row, exact=True)
    grade_columns = read_grade_columns(ws, header_row)

    table = {}
    for r in range(header_row + 1, ws.max_row + 1):
        level = cell_number(ws.cell(row=r, column=2).value)
        if level is None:
            break

        for grade, column in grade_columns.items():
            value = cell_number(ws.cell(row=r, column=column).value)
            if value is not None:
                table[(grade, int(level))] = int(round(value))

    return table


# ── 행 생성 ─────────────────────────────────────────────────────────

def make_index(category, grade, slot, level):
    """C(1) G(1) S(1) LL(2) 자리를 조합한 고정 Index. 사람이 읽을 수 있고 충돌하지 않는다."""
    grade_code = [g for _, g in GRADES].index(grade)
    return (CATEGORY_CODE[category] * 100000
            + grade_code * 10000
            + SLOT_CODE[slot] * 1000
            + level)


def build_rows(category, stat_table, shop_table, variants, main_stat):
    items, equipment, missing_price = [], [], []

    grade_ko_of = {en: ko for ko, en in GRADES}

    for (grade, level), base in sorted(stat_table.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        purchase = shop_table.get((grade, level))
        sell_price = int(round(purchase * SELL_PRICE_RATE)) if purchase is not None else 0
        if purchase is None:
            missing_price.append(f"{category} {grade} Lv{level}")

        option_count = OPTION_COUNT_OVERRIDE.get(
            (category, grade, level), OPTION_COUNT[grade])

        for slot, variant_ko in variants:
            index = make_index(category, grade, slot, level)

            items.append({
                "Index": index,
                "Name": NAME_TEMPLATE.format(
                    grade_ko=grade_ko_of[grade], variant_ko=variant_ko, level=level),
                "Category": category,
                "Grade": grade,
                "Level": level,
                "IconAddress": ICON_TEMPLATE.format(
                    category=category, slot=slot, grade=grade),
                "MaxStack": 1,
                "SellPrice": sell_price,
                "Description": ""
            })

            equipment.append({
                "Index": index,
                "EquipSlot": "Weapon" if category == "Weapon" else slot,
                "WeaponType": slot if category == "Weapon" else "None",
                "MainStatType": main_stat,
                "MainStatBase": base,
                "OptionCount": option_count
            })

    return items, equipment, missing_price


def write_json(filename, rows):
    path = os.path.join(OUTPUT_DIR, filename)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(f"  {filename}: {len(rows)}행 → {path}")


# ── 진입점 ──────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        sys.exit(__doc__)

    xlsx_path = sys.argv[1]
    if not os.path.isfile(xlsx_path):
        sys.exit(f"파일을 찾을 수 없습니다: {xlsx_path}")

    if not os.path.isdir(OUTPUT_DIR):
        sys.exit(f"출력 폴더가 없습니다: {OUTPUT_DIR}")

    workbook = openpyxl.load_workbook(xlsx_path, data_only=True)

    print("시트 읽는 중...")
    weapon_stats = read_stat_table(workbook["무기 수치표"])
    armor_stats = read_stat_table(workbook["방어구 수치표"])

    currency = workbook["재화 설계"]
    weapon_shop = read_shop_table(currency, "장비 상점 — 무기")
    armor_shop = read_shop_table(currency, "장비 상점 — 방어구")

    weapon_items, weapon_equipment, weapon_missing = build_rows(
        "Weapon", weapon_stats, weapon_shop, WEAPON_VARIANTS, "AttackDamage")
    armor_items, armor_equipment, armor_missing = build_rows(
        "Armor", armor_stats, armor_shop, ARMOR_VARIANTS, "Defense")

    items = sorted(weapon_items + armor_items, key=lambda row: row["Index"])
    equipment = sorted(weapon_equipment + armor_equipment, key=lambda row: row["Index"])

    print("\n생성:")
    write_json("ItemData.json", items)
    write_json("EquipmentData.json", equipment)

    # 판매가가 없는 조합은 SellPrice 0으로 나간다. 조용히 0이 되면 나중에 원인을 못 찾으므로 반드시 알린다.
    missing = weapon_missing + armor_missing
    if missing:
        print(f"\n[확인 필요] 상점 가격이 없어 SellPrice=0으로 처리한 조합 {len(missing)}건:")
        for entry in missing:
            print(f"  - {entry}")

    print(f"\n완료. 장비 {len(items)}종.")
    print("Unity에서 두 JSON을 Addressable로 등록하고 주소를 'ItemData' / 'EquipmentData'로 지정하세요.")


if __name__ == "__main__":
    main()
